"""
GrabFood Menu Scraper
Restaurant : Ayam Katsu Katsunami - Lokarasa Citraland
URL        : https://food.grab.com/id/id/restaurant/ayam-katsu-katsunami-lokarasa-citraland-delivery/6-C7EYGBJDME3JRN

Approach   : Intercept Grab's internal API via Playwright network listener.
             GrabFood is a React SPA — static HTTP requests return empty DOM.
             The real data lives in: GET /v5/merchant/get?merchantID=...
"""

import asyncio
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from playwright.async_api import async_playwright

TARGET_URL = (
    "https://food.grab.com/id/id/restaurant/"
    "ayam-katsu-katsunami-lokarasa-citraland-delivery/6-C7EYGBJDME3JRN"
)
MERCHANT_ID   = "6-C7EYGBJDME3JRN"
OUTPUT_FILE   = "katsunami_menu_data.xlsx"

# ─────────────────────────────────────────────────────────────
# NETWORK INTERCEPTION
# ─────────────────────────────────────────────────────────────

captured_data = {}

async def handle_response(response):
    """Capture the merchant API response that contains menu data."""
    url = response.url
    if MERCHANT_ID in url and ("merchant" in url or "merchants" in url or "get" in url):
        # Filter out images/static assets
        if "json" not in response.headers.get("content-type", "") and not url.endswith((".json", ".graphql")):
            # Fallback if content-type isn't parsed yet
            if any(ext in url for ext in [".jpg", ".png", ".webp", ".css", ".js", ".woff"]):
                return
                
        try:
            body = await response.json()
            # Verify it actually has merchant data
            if "merchant" in body or ("data" in body and "merchant" in body["data"]):
                captured_data["merchant"] = body
                print(f"[✓] Captured API response from: {url[:80]}...")
        except Exception:
            pass


async def scrape():
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"]
        )

        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="id-ID",
            viewport={"width": 1280, "height": 900},
        )

        page = await context.new_page()
        page.on("response", handle_response)

        print(f"[→] Opening {TARGET_URL}")
        await page.goto(TARGET_URL, wait_until="networkidle", timeout=60_000)
        await asyncio.sleep(3)  # allow lazy-loaded sections to fire

        await browser.close()

    return captured_data.get("merchant")


# ─────────────────────────────────────────────────────────────
# PARSER
# ─────────────────────────────────────────────────────────────

def parse_menu(raw: dict) -> list[dict]:
    """
    Extract menu items from Grab API JSON.
    Structure: data.menu.sections[].items[]
    (exact key path may vary — adjust if Grab changes their schema)
    """
    rows = []

    # Debug dump to see the schema
    with open("grab_api_debug.json", "w") as f:
        json.dump(raw, f, indent=2)

    # Try common paths
    merchant_data = raw.get("merchant", raw)
    
    outlet_name = merchant_data.get("name", "Ayam Katsu Katsunami - Lokarasa Citraland")
    
    menu_root = merchant_data.get("menu", raw)
    
    sections = menu_root.get("categories", menu_root.get("sections", []))
    if not sections:
        # fallback: look for any key that holds a list
        for v in menu_root.values():
            if isinstance(v, list) and len(v) > 0:
                sections = v
                break

    for section in sections:
        category = section.get("name", "Uncategorized")
        items    = section.get("items", [])

        for item in items:
            name = item.get("name", "")
            description = item.get("description", "")
            
            # ── prices ──────────────────────────────────────────
            price_before = 0
            price_after = 0
            promo_label = "-"

            # Original price
            if "priceInMinorUnit" in item:
                price_before = item.get("priceInMinorUnit", 0) / 100
                
            # Promos
            if "discountedPriceInMin" in item:
                price_after = item.get("discountedPriceInMin", 0) / 100
                
                # Check for promo label
                promo_label = item.get("discountPercentage") or item.get("campaignName") or "Promo"
                if not promo_label and price_before > price_after:
                    promo_label = f"Rp {int(price_before - price_after):,} OFF"
            else:
                price_after = price_before

            # ── availability ─────────────────────────────────────
            available_raw = item.get("available", True)
            available     = "Tersedia" if available_raw else "Habis"

            rows.append({
                "outlet":       outlet_name,
                "category":     category,
                "name":         name,
                "description":  description,
                "price_before": int(price_before),
                "price_after":  int(price_after),
                "promo":        promo_label,
                "available":    available,
            })

    return rows, outlet_name


# ─────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────

def export_excel(rows: list[dict], outlet_name: str, filename: str):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Menu Data"

    HDR_BG    = "1A1A2E"
    HDR_FG    = "FFFFFF"
    PROMO_BG  = "D4EDDA"
    HABIS_BG  = "F8D7DA"
    ODD_BG    = "F8F9FA"
    EVEN_BG   = "FFFFFF"

    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── title ──
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = f"DATA MENU — {outlet_name.upper()}"
    c.font      = Font(name="Calibri", bold=True, size=13, color=HDR_FG)
    c.fill      = PatternFill("solid", fgColor=HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    c2 = ws["A2"]
    c2.value     = f"Scraped: {datetime.now().strftime('%d %B %Y, %H:%M WIB')}  |  Source: GrabFood  |  ID: {MERCHANT_ID}"
    c2.font      = Font(name="Calibri", italic=True, size=9, color="888888")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    # ── headers ──
    headers   = ["No","Nama Outlet","Kategori","Nama Menu","Deskripsi",
                 "Harga Sebelum Promo","Harga Setelah Promo","Promo","Ketersediaan"]
    col_widths = [5, 28, 18, 30, 45, 20, 20, 16, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=HDR_FG)
        cell.fill      = PatternFill("solid", fgColor=HDR_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 22

    # ── data rows ──
    for i, row in enumerate(rows, 1):
        r   = 4 + i
        promo = row["promo"] != "-"
        habis = row["available"] == "Habis"
        bg    = HABIS_BG if habis else (PROMO_BG if promo else (ODD_BG if i % 2 else EVEN_BG))
        fill  = PatternFill("solid", fgColor=bg)

        vals = [i, row["outlet"], row["category"], row["name"], row["description"],
                row["price_before"], row["price_after"], row["promo"], row["available"]]

        for ci, val in enumerate(vals, 1):
            cell            = ws.cell(row=r, column=ci, value=val)
            cell.fill       = fill
            cell.border     = border
            cell.alignment  = Alignment(vertical="center", wrap_text=True,
                                        horizontal="center" if ci in [1,6,7,8,9] else "left")
            if ci in [6, 7]:
                cell.number_format = '"Rp "#,##0'
            bold_green  = ci == 8 and promo
            bold_red    = ci == 9 and habis
            cell.font   = Font(name="Calibri", size=10,
                               bold=bold_green or bold_red or (ci == 7 and promo),
                               color=("155724" if bold_green else ("721C24" if bold_red else "000000")))
        ws.row_dimensions[r].height = 38

    ws.freeze_panes = "A5"
    wb.save(filename)
    print(f"[✓] Saved: {filename}  ({len(rows)} rows)")


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

async def main():
    print("=" * 60)
    print("GrabFood Menu Scraper — Katsunami Lokarasa Citraland")
    print("=" * 60)

    raw = await scrape()

    if not raw:
        print("\n[!] No API data captured. Possible reasons:")
        print("    • Anti-bot / rate limit triggered")
        print("    • Grab changed the API endpoint path")
        print("    • Network timeout")
        print("\n    Try: add playwright-stealth, rotate user-agent, increase timeout.")
        return

    rows, outlet_name = parse_menu(raw)

    if not rows:
        print("[!] Parsed 0 items. Check parse_menu() key paths.")
        return

    export_excel(rows, outlet_name, OUTPUT_FILE)
    print(f"\n[✓] Done. Output: {OUTPUT_FILE}")


if __name__ == "__main__":
    asyncio.run(main())
